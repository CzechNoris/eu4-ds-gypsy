#!/usr/bin/env python

import os
import pandas as pd
from openpyxl.styles import PatternFill

import argparse
import logging


import sys
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
from src.loader_tools import load_ideas, load_policies
from src.model import Idea, Policy

def setup_logging(log_level):
    numeric_level = getattr(logging, log_level.upper(), None)
    if not isinstance(numeric_level, int):
        raise ValueError('Invalid log level: %s' % log_level)

    logging.basicConfig(level=numeric_level,
                        format='%(levelname)s %(message)s')


parser = argparse.ArgumentParser(
    description='Process EU4 policy files into YAML format.')
parser.add_argument('--log-level', dest='log_level', default='WARN',
                    choices=['DEBUG', 'INFO', 'WARN', 'ERROR', 'CRITICAL'],
                    help='Set the logging level (default: warning)')
parser.add_argument('-i', '--idea_files', nargs='*',
                    required=True, help='List of idea files for the table.')
parser.add_argument('-p', '--policy_files', nargs='*',
                    required=True, help='List of policy files for the table.')
parser.add_argument('-o', '--output_folder_path', type=str, default='export/',
                    help='Path to output folder.')
args = parser.parse_args()
setup_logging(args.log_level)

if 'export/' not in args.output_folder_path:
    logging.warning('Output folder should be under export/!')

if not os.path.exists(args.output_folder_path):
    logging.info('Creating output folder: %s', args.output_folder_path)
    os.makedirs(args.output_folder_path)

for file_path in args.idea_files + args.policy_files:
    if not os.path.exists(file_path):
        logging.error('File does not exist: %s', file_path)
        exit(1)

IDEAS: dict[str, Idea] = load_ideas(args.idea_files)
POLICIES: dict[str, Policy] = load_policies(args.policy_files)

# TODO: make this constants loaded from files
RELIGIOUS_IDEA_NAMES = {'religious_ideas', 'anglican0', 'animist0', 'buddhism0', 'cathar0', 'catholic0', 'confucian0', 'coptic0', 'dreamtime0', 'fetishist0', 'hellenic0', 'hindu0', 'hussite0', 'ibadi0', 'inti0',
                        'jewish0', 'manichean0', 'mesoamerican0', 'nahuatl0', 'norse0', 'orthodox0', 'protestant0', 'reformed0', 'romuva0', 'shia0', 'shinto0', 'slavic0', 'sunni0', 'suomi0', 'tengri0', 'totemist0', 'zoroastrian0'}
GOVERNMENT_IDEA_NAMES = {'dictatorship0', 'horde0',
                         'monarchy0', 'republic0', 'theocracy0'}


def get_idea_type(idea_name: str) -> str:
    if idea_name in IDEAS:
        return IDEAS[idea_name].type
    elif idea_name == 'religious_ideas':
        return 'ADM'
    elif idea_name == 'government_ideas':
        return 'ADM'
    else:
        raise ValueError(f'Unknown idea name: {idea_name}')


GROUPED_IDEA_NAMES = [idea.name for idea in IDEAS.values(
) if idea.name not in RELIGIOUS_IDEA_NAMES and idea.name not in GOVERNMENT_IDEA_NAMES] + ['religious_ideas', 'government_ideas']
GROUPED_IDEA_NAMES = sorted(GROUPED_IDEA_NAMES)
GROUPED_IDEA_NAMES = sorted(
    GROUPED_IDEA_NAMES, key=lambda x: (get_idea_type(x), x))


def replace_idea_name(idea_name):
    if idea_name in RELIGIOUS_IDEA_NAMES:
        return 'religious_ideas'
    elif idea_name in GOVERNMENT_IDEA_NAMES:
        return 'government_ideas'
    else:
        return idea_name


policy_table = pd.DataFrame(
    index=GROUPED_IDEA_NAMES,
    columns=GROUPED_IDEA_NAMES,
)
policy_table = policy_table.fillna('')

for policy in POLICIES.values():
    req1, req2 = policy.req
    policy_text = f'{policy.type} {policy.name}\n'
    for key, value in policy.effect.items():
        policy_text += f'{key}: {value}\n'
    for idea_a_name in req1:
        for idea_b_name in req2:
            idea_a_name = replace_idea_name(idea_a_name)
            idea_b_name = replace_idea_name(idea_b_name)
            if not policy_table.loc[idea_a_name, idea_b_name]:
                policy_table.loc[idea_a_name, idea_b_name] = policy_text
                policy_table.loc[idea_b_name, idea_a_name] = policy_text
            else:
                if policy_table.loc[idea_a_name, idea_b_name] != policy_text:
                    logging.warning(
                        f'Conflict: {policy_table.loc[idea_a_name, idea_b_name].split()[1]} vs {policy.name} for {idea_a_name} and {idea_b_name}')

with pd.ExcelWriter(os.path.join(args.output_folder_path, 'policy_table.xlsx'), engine="openpyxl") as writer:
    sheet_name = 'policy_table'
    policy_table.to_excel(writer, sheet_name=sheet_name,
                          index=True, header=True)
    sheet = writer.sheets[sheet_name]

    column_range = sheet['B:CA']

    for column in column_range:
        for cell in column[1:100]:
            if cell.value:
                if cell.value.startswith('adm') or cell.value.startswith('ADM'):
                    cell.fill = PatternFill("solid", fgColor="5cb800")
                elif cell.value.startswith('dip') or cell.value.startswith('DIP'):
                    cell.fill = PatternFill("solid", fgColor="00b0f0")
                elif cell.value.startswith('mil') or cell.value.startswith('MIL'):
                    cell.fill = PatternFill("solid", fgColor="ff2800")
            
